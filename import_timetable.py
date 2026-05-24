from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
PERIOD_TIMES_PATH = ROOT / "period_times.json"
DEFAULT_OUTPUT_PATH = ROOT / "courses.csv"

WEEKDAY_MAP = {
    "星期一": "1",
    "周一": "1",
    "星期二": "2",
    "周二": "2",
    "星期三": "3",
    "周三": "3",
    "星期四": "4",
    "周四": "4",
    "星期五": "5",
    "周五": "5",
    "星期六": "6",
    "周六": "6",
    "星期日": "7",
    "星期天": "7",
    "周日": "7",
    "周天": "7",
}

FIELDNAMES = [
    "semester_id",
    "course_id",
    "title",
    "weekday",
    "start_time",
    "end_time",
    "start_week",
    "end_week",
    "week_type",
    "location",
    "teacher",
    "notes",
    "alarm_minutes",
]


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def load_period_times(path: Path) -> dict[str, dict[str, str]]:
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    periods = data.get("periods", {})
    if not isinstance(periods, dict):
        raise ValueError("period_times.json must contain a periods object")
    return periods


def parse_period_range(period_text: str) -> tuple[int, int]:
    text = clean(period_text)
    match = re.search(r"(\d+)\s*(?:-|~|—|至)\s*(\d+)", text)
    if match:
        return int(match.group(1)), int(match.group(2))
    match = re.search(r"\d+", text)
    if match:
        period = int(match.group(0))
        return period, period
    raise ValueError(f"cannot parse period range from {period_text!r}")


def parse_weeks(week_text: str) -> list[tuple[int, int]]:
    text = clean(week_text)
    text = re.sub(r"\[[^\]]+\]", "", text)
    text = text.replace("周", "").replace("，", ",").replace("、", ",")
    text = re.sub(r"\s+", "", text)
    parts = [part for part in text.split(",") if part]
    ranges: list[tuple[int, int]] = []
    for part in parts:
        match = re.fullmatch(r"(\d+)(?:-|~|—|至)(\d+)", part)
        if match:
            start, end = int(match.group(1)), int(match.group(2))
        else:
            start = end = int(part)
        if end < start:
            raise ValueError(f"week range end before start in {week_text!r}")
        ranges.append((start, end))
    if not ranges:
        raise ValueError(f"cannot parse weeks from {week_text!r}")
    return ranges


def detect_week_type(text: str) -> str:
    if "单周" in text or clean(text) == "单周":
        return "odd"
    if "双周" in text or clean(text) == "双周":
        return "even"
    return "all"


def extract_period_text(week_text: str, section_text: str) -> str:
    bracket = re.search(r"\[([^\]]+)\]", clean(week_text))
    if bracket:
        return bracket.group(1)
    return section_text


def build_notes(row: dict[str, str], source_row: int, week_text: str, section_text: str) -> str:
    parts = [f"源表第 {source_row} 行", f"原周次: {week_text}", f"原节次: {section_text}"]
    if row.get("授课班级"):
        parts.append(f"班级: {row['授课班级']}")
    if row.get("上课人数"):
        parts.append(f"人数: {row['上课人数']}")
    if row.get("学时分类"):
        parts.append(f"学时分类: {row['学时分类']}")
    if row.get("平台信息"):
        parts.append(f"平台信息: {row['平台信息']}")
    return "；".join(parts)


def sheet_rows(path: Path, sheet_name: str | None) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    values = list(worksheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = [clean(value) for value in values[0]]
    rows = []
    current_weekday = ""
    for offset, row_values in enumerate(values[1:], start=2):
        row = {headers[index]: clean(value) for index, value in enumerate(row_values) if index < len(headers)}
        if row.get("星期"):
            current_weekday = row["星期"]
        row["星期"] = current_weekday
        if any(row.values()):
            row["_source_row"] = str(offset)
            rows.append(row)
    return rows


def convert_rows(
    rows: list[dict[str, str]],
    periods: dict[str, dict[str, str]],
    semester_id: str,
) -> list[dict[str, str]]:
    output_rows: list[dict[str, str]] = []
    for row in rows:
        source_row = int(row["_source_row"])
        weekday_text = clean(row.get("星期"))
        if weekday_text not in WEEKDAY_MAP:
            raise ValueError(f"row {source_row}: unknown weekday {weekday_text!r}")

        title = clean(row.get("课程名称"))
        week_text = clean(row.get("周次"))
        section_text = clean(row.get("节次"))
        if not title or not week_text:
            continue

        period_text = extract_period_text(week_text, section_text)
        start_period, end_period = parse_period_range(period_text)
        if str(start_period) not in periods or str(end_period) not in periods:
            raise ValueError(f"row {source_row}: period {start_period}-{end_period} missing from period_times.json")

        week_type = detect_week_type(f"{week_text} {section_text}")
        notes = build_notes(row, source_row, week_text, section_text)
        for start_week, end_week in parse_weeks(week_text):
            output_rows.append(
                {
                    "semester_id": semester_id,
                    "course_id": f"r{source_row:02d}",
                    "title": title,
                    "weekday": WEEKDAY_MAP[weekday_text],
                    "start_time": periods[str(start_period)]["start"],
                    "end_time": periods[str(end_period)]["end"],
                    "start_week": str(start_week),
                    "end_week": str(end_week),
                    "week_type": week_type,
                    "location": clean(row.get("上课地点")),
                    "teacher": clean(row.get("任课教师")),
                    "notes": notes,
                    "alarm_minutes": "",
                }
            )
    return output_rows


def read_existing_courses(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        return [{field: clean(row.get(field)) for field in FIELDNAMES} for row in reader]


def write_courses(path: Path, rows: list[dict[str, str]], append: bool = False) -> None:
    output_rows = [*read_existing_courses(path), *rows] if append else rows
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(output_rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Import a one-dimensional timetable .xlsx into courses.csv.")
    parser.add_argument("--input", required=True, type=Path, help="Path to the timetable .xlsx file.")
    parser.add_argument("--sheet", default=None, help="Worksheet name. Defaults to the first sheet.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT_PATH, type=Path, help="Output courses.csv path.")
    parser.add_argument("--semester-id", default="default", help="Semester id from semesters.csv.")
    parser.add_argument("--append", action="store_true", help="Append imported rows to the existing output file.")
    args = parser.parse_args()

    periods = load_period_times(PERIOD_TIMES_PATH)
    rows = convert_rows(sheet_rows(args.input, args.sheet), periods, args.semester_id)
    write_courses(args.output, rows, append=args.append)
    mode = "appended to" if args.append else "imported into"
    print(f"Imported {len(rows)} course rows {mode} {args.output}.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        raise SystemExit(1)
